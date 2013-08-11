'''
Log Data - dialog for logging workouts

Dependencies:
    Python 2.7.5
    Tkinter 8.5.3

Author: Tyler Weaver

Revision History:
(20 June 2013)
    Logger class
(25 June 2013)
    Initial GUI design
(30 June 2013)
    Remove email field for middle name
(01 July 2013)
    Update customer list after new customer has been added.
    Accommodated for suffixes (Jr, III, etc.)
(02 July 2013)
    Add date field
(03 July 2013)
    Disable/Enable date text field
    Fixed set workout bug

further revision history maintained in git
    
'''

from Tkinter import StringVar,E,W,Toplevel
from ttk import Frame, Button, Entry, Label, Combobox
from ScrolledText import ScrolledText
from tkMessageBox import showerror,askquestion
from DialogTemplate import Dialog
from datetime import datetime,date,timedelta,time
from time import strftime

from customer import NewCustomerDialog, Customers
from schedule import Schedule
from payment import Payments
from reports import workouts_this_month

from openpyxl import Workbook, load_workbook
from openpyxl.style import Border, Fill
from openpyxl.shared.exc import InvalidFileException

from pprint import pprint

class LoggerDialog(Toplevel):
    def __init__(self, master, customers, payments):
        Toplevel.__init__(self,master)

        self.root = master

        self.title("Check In")
        self.iconname = "Check In"

        self.name = StringVar() # variable for customer
        self.customers = customers # customers object
        self.payments = payments
        self.names = []
        self.workout = StringVar()
        self.workouts = []
        self.workouts_form = []
        self.date = StringVar()
        self.date.set(strftime("%m/%d/%Y"))
        self.refresh_time = 15 # in minutes
        self.output = '' # for the output label at the bottom
        self.schedule = Schedule()

        self.logger = Logger() #throws IOError if file is open

        inf = Frame(self)
        inf.pack(padx=10,pady=10,side='top')
        Label(inf, text="Name:").grid(row=0,column=0,sticky=E,ipady=2,pady=2,padx=10)
        Label(inf, text='Date:').grid(row=1,column=0,sticky=E,ipady=2,pady=2,padx=10)
        Label(inf, text="Workout:").grid(row=2,column=0,sticky=E,ipady=2,pady=2,padx=10)

        self.name_cb = Combobox(inf, textvariable=self.name, width=30,
                                values=self.names)
        self.name_cb.grid(row=0,column=1,sticky=W,columnspan=2)
        self.date_ent = Entry(inf, textvariable=self.date)
        self.date_ent.grid(row=1,column=1,sticky=W)
        self.date_ent.bind('<FocusOut>', self.update_workouts)
        Button(inf,text='Edit', command=self.enable_date_ent).grid(row=1,column=2,sticky=E)
        self.workout_cb = Combobox(inf, textvariable=self.workout, width=30,
                                   values=self.workouts_form,state='readonly')
        self.workout_cb.grid(row=2,column=1,sticky=W,columnspan=2)

        self.log_btn=Button(inf,text="Log Workout",command=self.log,width=12)
        self.log_btn.grid(row=3,column=1,columnspan=2,pady=4,sticky='ew')
        
        stf = Frame(self)
        stf.pack(padx=10,pady=10,fill='x',side='top')
        self.scrolled_text = ScrolledText(stf,height=15,width=50,wrap='word',state='disabled')
        self.scrolled_text.pack(expand=True,fill='both')

        self.update_workouts()
        self.update_names()

        self.bind('<Return>',self.log)
        self.name_cb.focus_set()  # set the focus here when created

        #disable the date field
        self.disable_date_ent()

        #start time caller
        self.time_caller()

    def output_text(self,outstr):
        self.scrolled_text['state'] = 'normal'
        self.scrolled_text.insert('end',outstr)
        self.scrolled_text.see('end')
        self.scrolled_text['state'] = 'disabled'

    def log(self, e=None):
        #check to see if name is blank
        logged = False
        if self.name.get() == '':
            self.output_text("! - Please select your name.\n")
        elif self.workout.get() not in self.workouts_form:
            self.output_text("! - Select valid workout.\n")
        elif self.name.get() not in self.names: # new customer
            self.new_customer_error()
        else: # log the workout
            name = self.name.get().split(' ',2)
            (line, r) = self.customers.find(name[2],name[0],name[1])
            name_str = str(self.name.get())
            date = datetime.strptime(str(self.date.get()),'%m/%d/%Y')

            if not line:
                self.output_text("!! - No record: " + self.name.get() + ".\n")

            while (not logged):
                try:
                    self.logger.log(self.workouts[self.workout_cb.current()][0],
                                    self.workouts[self.workout_cb.current()][1],
                                    name_str, day=date)
                    logged = True
                except IOError:
                    showerror("Error writting to file", "Please close " + self.logger.filename + " and press OK.")


            if logged:
                self.output_text(self.name.get() + " - " + line[3] + "\n")
                logged_payment = False
                while(not logged_payment):
                    try:
                        if line[3] == 'Monthly':
                            if not self.payments.has_paid_monthly(name_str):
                                self.output_text("$ - Please pay your monthly dues.\n")
                        elif line[3] == 'Punch Card':
                            punch = self.payments.punch(name_str)
                            if not punch:
                                self.output_text("$ - Please purchase another punch card.\n")
                            else:
                                self.output_text("$ - You have " + str(punch) + " remaining workouts on your card.\n")
                        elif line[3] == 'Drop In':
                            self.payments.drop_in(name_str, date)
                            self.output_text("$ - Drop In payment logged.\n")
                        logged_payment = True
                    except IOError:
                        # this is bad, you logged a workout and you failed to log payment
                        showerror("Error writting to file", "Please close " + self.payments.filename + " and press OK.")
                    else:
                        #exception not raised
                        try: #accessing log file here
                            workout_count = str(workouts_this_month(name_str,self.logger.filename,date.strftime("%B"))) 
                            self.output_text("Workouts you've completed this month: " + workout_count + "\n")
                        except IOError:
                            showerror("Error reading from file", "Please close " + self.logger.filename + " and press OK.")


                self.update_time_now()
                self.set_workout_now()
                self.update_workouts()
            
    def new_customer_error(self):
        self.ncd = NewCustomerDialog(self,self.customers)
        if askquestion(title="New Customer?",
            message="Add new customer: " + self.name.get(),
            parent = self) == 'yes':

            temp = self.name.get().split(' ')
            self.ncd.fname.set(temp[0])
            if len(temp) == 2:
                self.ncd.lname.set(temp[1])
            elif len(temp) == 3:
                self.ncd.mname.set(temp[1])
                self.ncd.lname.set(temp[2])
            elif len(temp) > 3:
                self.ncd.mname.set(temp[1])
                self.ncd.lname.set(' '.join(temp[2:4]))

            self.ncd.show()

        if self.ncd.new_customer_name:
            self.add_name(self.ncd.new_customer_name)
            self.output_text("+ - " + self.ncd.new_customer_name + " added.\n")

    def disable_date_ent(self, e=None):
        self.date_ent['state'] = 'disabled'

    def enable_date_ent(self, e=None):
        self.date_ent['state'] = 'normal'
        
    def time_caller(self):
        #updates every 15 min automatically
        msec = self.refresh_time * 60 * 100

        self.update_time_now() #update time to current time
        self.set_workout_now()
        self.update_workouts() #update the workouts
        
        self.after(msec, self.time_caller) #call again

    def update_time_now(self):
        self.enable_date_ent()
        self.date.set(strftime("%m/%d/%Y"))

    def set_workout_now(self):
        #set workout field
        if len(self.workouts) == 0:
            self.disable_date_ent()
            return #no workouts
        index = 0
        now = datetime.today()
        for i, workout in enumerate(self.workouts):
            test = datetime.combine(date.today(),workout[0])
            if now < (test - timedelta(minutes=15)):
                index = i
                break
        self.workout_cb.current(index)
        self.disable_date_ent()
            
    def update_workouts(self, e=None):
        try:
            self.populate_workouts()
            self.workout_cb['values'] = []
            self.workout_cb['values'] = self.workouts_form
        except ValueError:
            self.workout.set(' Enter Valid Date ')
        if len(self.workouts) > 0 and e:
            self.workout_cb.current(0)
            
    def populate_workouts(self):
        today = datetime.strptime(str(self.date.get()), "%m/%d/%Y") #get date
        dow = self.schedule.weekday_to_str(today.weekday()) #get dow string

        self.workouts = self.schedule.get_wkday(dow)
        self.workouts_form = []
        for w in self.workouts:
            self.workouts_form.append(w[0].strftime("%H:%M") + ' - ' + w[1])
        if len(self.workouts) == 0:
            self.workout.set(' No workouts today ')

    def update_names(self):
        self.populate_names()
        self.name_cb['values'] = self.names

    def add_name(self, name):
        self.names.append(name)
        split_names = [x.split(' ') for x in self.names]
        split_names.sort(key = lambda x: ' '.join([x[2],x[0],x[1]]))
        self.names = [' '.join(x) for x in split_names]
        self.name_cb['values'] = self.names
        self.name.set(name)
        
    def populate_names(self):
        try:
            clist = self.customers.get_list()
        except IOError:
            self.output_text("! - " + self.customers.filename + " open in another application.\n")
            return
        clist.sort(key = lambda x: ', '.join(x[0:3]).lower())
        self.names = []
        for line in clist:
            self.names.append(' '.join([line[1],line[2],line[0]]))

    def find_line(self, name):
        [fname, mname, lname] = name.split(' ')
        try:
            return self.customers.find(lname, fname, mname)
        except IOError:
            self.output_text("! - " + self.customers.filename + " open in another application.\n")
            return None

class Logger:
    def __init__(self):
        self.month = strftime("%B")
        self.year = strftime("%Y")
        self.filename = 'jim_data' + self.year + '.xlsx'
        try:
            self.wb = load_workbook(self.filename)
        except InvalidFileException:
            self.wb = Workbook()
            sh = self.wb.get_sheet_by_name("Sheet")
            self.wb.remove_sheet(sh)

        self.sh = self.wb.get_sheet_by_name(self.month)
        if not self.sh:
            self.wb.create_sheet(title=self.month)
            # print "Created new month log: " + self.month
            self.sh = self.wb.get_sheet_by_name(self.month)
            self.sh.append(['Date','Time','Class Type','Customer'])
            for col in range(4):
                cell = self.sh.cell(row=0, column=col).style
                cell.fill.fill_type = Fill.FILL_SOLID
                cell.fill.start_color.index = "FFDDD9C4"
                cell.borders.bottom.border_style = Border.BORDER_THIN

        self.sh.garbage_collect()
        self.wb.save(self.filename)


    def log(self, hour, class_type, customer, day=date.today()):
        line = [day, hour.strftime('%H:%M'), class_type, customer]
        self.sh.garbage_collect()
        self.sh.append(line)
        self.sh.cell(row=(self.sh.get_highest_row()-1),
                         column=0).style.number_format.format_code='m/d/yyyy'
        self.wb.save(self.filename)

class CheckInFrame(Frame):
    def __init__(self,master,customers,payments):
        Frame.__init__(self,master)
        self.customers = customers
        self.payments = payments
        self.master = master

        btn = Button(self,text="Open Check In Dialog",command=self.logger_diag,
            width=40)
        btn.pack(padx=100,pady=50)
        
    def logger_diag(self):
        LoggerDialog(self.master, self.customers, self.payments)

if __name__ == '__main__':
    root = Frame()
    root.pack()
    c = Customers()
    p = Payments()
    cif = CheckInFrame(root, c, p)
    cif.pack()
    root.mainloop()