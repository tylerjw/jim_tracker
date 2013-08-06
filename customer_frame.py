'''
Customer libraires for jim tracker

Author: Tyler Weaver

Dependencies:
    Python 2.7.5
    Tkinter 8.5.2

Revision history:
(25 June 2013)
    Clear values each time the dialog box is shown.
    Return bound to submit.
    Error box validation.
    Fixed unicode/string issue. Converted all to string when reading.
    Duplicate entry not allowed, button removed.
    Default focus set on first name entry
(30 June 2013)
    Remove email field for middle name
    Adjusted reads and writes to use str(val).strip()
    
TODO:
    Calender date select
    Red highlighted field validation instead of error dialog boxes.

    frame version
    requires packing
'''

from openpyxl import Workbook, load_workbook
from datetime import date, datetime
from Tkinter import Toplevel, StringVar, IntVar, W, E, LEFT, Message, BOTH
from ttk import Entry, Button, Frame, Label, Combobox,LabelFrame
from pprint import pprint
from DialogTemplate import Dialog
from tkMessageBox import showerror
import re

class NewCustomerFrame(LabelFrame):
    def __init__(self, master, customers):
        LabelFrame.__init__(self, master, text="New Customer")
        self.customers = customers
        self.fname = StringVar()
        self.lname = StringVar()
        self.mname = StringVar()
        self.payment = StringVar()
        self.date = StringVar()
        self.iconname="New Customer"

        # self.root = Toplevel(master)
        
        ### dialog content        
        Label(self, text="Name: ").grid(row=0,sticky='e',ipady=2,pady=2)
        Label(self, text="Type: ").grid(row=1,sticky='e',ipady=2,pady=2)
        Label(self, text="Date:").grid(row=1,column=3,sticky='e',ipady=2,pady=2)

        self.fname_en = Entry(self, width=20, textvariable=self.fname)
        self.mname_en = Entry(self, width=4, textvariable=self.mname)
        self.lname_en = Entry(self, width=20, textvariable=self.lname)
        self.payment_cb = Combobox(self, textvariable=self.payment, width=12,
                                   values=("Drop In", "Punch Card", "Monthly"))
        self.date_en = Entry(self, width=15, textvariable=self.date)

        Frame(self, width=5).grid(row=0,column=1,sticky=W)
        
        self.fname_en.grid(row=0,column=2,sticky=W)
        self.mname_en.grid(row=0,column=3,sticky='ew')
        self.lname_en.grid(row=0,column=4,sticky=W)
        self.payment_cb.grid(row=1,column=2,columnspan=2,sticky=W)
        self.date_en.grid(row=1,column=4,columnspan=2,sticky=W)
        
        ### buttons
        Button(self, text='Reset Values', width=15,
               command=self.reset_values).grid(row=3,column=0,columnspan=3,sticky='w',padx=10,pady=3)
        Button(self, text='Submit', width=15,
               command=self.add_customer).grid(row=3,column=4,sticky='e')

        # self.root.bind("<Return>", self.add_customer)
        self.fname_en.focus_set()

        self.reset_values()

    def reset_values(self):
        self.fname.set('')
        self.mname.set('')
        self.lname.set('')
        # blow out the field every time this is created
        self.date.set(date.today().strftime("%m/%d/%Y"))
        self.payment_cb.set("Drop In")

    def add_customer(self, event=None):
        # validate and show errors
        if self.fname.get() == '':
            showerror("Error!", "First name field blank!")
        elif self.lname.get() == '':
            showerror("Error!", "Last name field blank!")
        elif self.mname.get() == '':
            showerror("Error!", "Middle initial field blank!")
        elif self.payment.get() not in ("Drop In", "Punch Card", "Monthly"):
            showerror("Error!", "Incorect Payment type!")
        elif not re.compile(r'[01]?\d/[0123]?\d/[12]\d{1,3}').search(self.date.get()):
            showerror("Error!", "Bad entry for date, use format mm/dd/yyyy")
        else:
            # do work
            old, row = self.customers.find(str(self.lname.get()).strip(), str(self.fname.get()).strip(),
                                           str(self.mname.get()).strip())
            new = [str(self.lname.get()).strip(), str(self.fname.get()).strip(), str(self.mname.get()).strip(),
                   str(self.payment.get()).strip(), datetime.strptime(self.date.get(), "%m/%d/%Y")]
            
            if not old:
                self.customers.add(new)
            else:
                var = IntVar()
                
                diag = AlreadyExistsDialog(self, new, old, var)
                diag.show()
                if var.get() == 0: # edit
                    pass
                if var.get() == 1: # replace
                    self.customers.replace(row, new)
                # elif var.get() == 2: # add duplicate
                #     self.customers.add(new)
            
            # if close:             #this is a frame, we don't close any more
            #     self.quit()

class AlreadyExistsDialog(Dialog):
    def __init__(self, master_frame, new, old, variable, class_=None, relx=0.5, rely=0.3):
        Dialog.__init__(self, master_frame, 'Warning!',
                        class_, relx, rely)
        self.new = new
        self.old = old
        self.variable = variable
        
        self.text = "Warning: Customer Found with the same name.\n" + \
        "\nNew record: " + new[1] + " " + new[2] + " " + new[0] + \
        " (" + new[3] + ") - " + new[4].strftime("%m/%d/%Y") + \
        "\nOld record: " + old[1] + " " + old[2] + " " + old[0] + \
        " (" + old[3] + ") - " + old[4].strftime("%m/%d/%Y") + \
        "\n\n Cancel to edit entry, Override to replace old entry."
        

    def show(self):
        self.setup()

        self.msg = Message(self.root, text=self.text, aspect=400)
        self.msg.pack(expand=True, fill=BOTH)

        self.frame = Frame(self.root)
        
        b1 = Button(self.frame, text="Cancel", command=self.cancel)
        b1.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        b2 = Button(self.frame, text="Override", command=self.replace)
        b2.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        #b3 = Button(self.frame, text="Add Duplicate", command=self.add_duplicate)
        #b3.pack(side=LEFT, fill=BOTH, expand=True)
        
        self.frame.pack(padx=10, pady=10)

        self.root.bind('<Return>', self.cancel)
        
        self.enable()

    def cancel(self, event=None):
        self.variable.set(0)
        self.root.quit()

    def replace(self):
        self.variable.set(1)
        self.root.quit()

    def add_duplicate(self):
        self.variable.set(2)
        self.wm_delete_window()
        

class Customers:
    def __init__(self, filename='jim_info.xlsx', sheet_name='Customers'):
        self.wb = load_workbook(filename)
        self.filename = filename
        self.sh = self.wb.get_sheet_by_name(sheet_name)
        self.sheet_name = sheet_name
        if not self.sh:
            print "Error opening " + sh_name + " sheet."

    def reload_file(self):
        self.wb = load_workbook(self.filename)
        self.sh = self.wb.get_sheet_by_name(self.sheet_name)

    def add(self, new):
        ''' new = [lname, fname, email, ptype, date_created] '''
        self.sh.garbage_collect()
        # Append new customer
        self.sh.append(new)
        # set the date style
        self.sh.cell(row=(self.sh.get_highest_row()-1),
                     column=4).style.number_format.format_code = 'm/d/yyyy'    
        
        # save the file
        self.wb.save(self.filename)

    def replace(self, row, new):
        ''' new = [lname, fname, mname, ptype, date_created] '''
        for col in range(5):
            self.sh.cell(row=row,column=col).value = new[col]

        self.sh.cell(row=row,
                     column=4).style.number_format.format_code = 'm/d/yyyy'  
        # save the file
        self.wb.save(self.filename)

    def get_list(self):
        self.reload_file()
        output = []
        for r in range(1, self.sh.get_highest_row()):
            points = self.sh.range("A1:E1",row=r)
            values = []
            for i in range(5):
                values.append(str(points[0][i].value).strip())
            
            output.append(values)
        return output

    def get_dict(self):
        values = self.get_list()
        output = {}
        for r in range(len(values)):
            name = ', '.join(values[r][:2]).strip()
            output[name] = values[r][2:]
        return output
    
    def find(self, lname, fname, mname):
        self.reload_file()
        output = None
        row = None
        for r in range(self.sh.get_highest_row()):
            if (lname == str(self.sh.cell(row=r, column=0).value).strip() and
                fname == str(self.sh.cell(row=r, column=1).value).strip() and 
                mname == str(self.sh.cell(row=r, column=2).value).strip()):
                temp = self.sh.range("A1:E1",row=r)
                output = []
                for i in range(5):
                    if i != 4:
                        output.append(str(temp[0][i].value).strip())
                    else:
                        output.append(temp[0][i].value)
                row = r
                break
        return output, row

if __name__ == '__main__':
    c = Customers()
    pprint(c.get_list())
    
    ncf = NewCustomerFrame(None, c)
    ncf.mainloop()