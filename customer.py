'''
Customer libraires for jim tracker

Author: Tyler Weaver

Dependencies:
    Python 2.7.5
    Tkinter 8.5.2

Revision history:
(25 June 2013)
    Clear values each time the dialog box is shown.
    Return bound to submit.
    Error box validation.
    Fixed unicode/string issue. Converted all to string when reading.
    Duplicate entry not allowed, button removed.
    Default focus set on first name entry
(30 June 2013)
    Remove email field for middle name
    Adjusted reads and writes to use str(val).strip()
    
    see git repository for versioning
'''

from openpyxl import Workbook, load_workbook
from datetime import date, datetime
from Tkinter import Toplevel, StringVar, IntVar, W, E, LEFT, Message, BOTH, Tk
from ttk import Entry, Button, Frame, Label, Combobox,LabelFrame
from pprint import pprint
from DialogTemplate import Dialog
from tkMessageBox import showerror
import re


class NewCustomerDialog(Dialog):
    def __init__(self, master, customers, edit=False, class_=None, relx=0.5, rely=0.3):
        if edit:
            self.title = "Edit Customer"
        else:
            self.title = "New Customer"

        Dialog.__init__(self, master, self.title,
                        class_, relx, rely)
        self.customers = customers

        self.fname = StringVar()
        self.lname = StringVar()
        self.mname = StringVar()
        self.payment = StringVar()
        self.date = StringVar()
        self.close = False

        self.payment.set("Drop In")

        self.edit = edit

        self.new_customer_name = None

    def show(self, line=None):
        '''
        allows preset values
        '''
        self.setup() #base class setup

        self.frame = Frame(self.root)

        # blow out the field every time this is created
        if not self.edit: self.date.set(date.today().strftime("%m/%d/%Y"))
        
        ### dialog content        
        Label(self.frame, text="First Name: ").grid(row=0, sticky=W, ipady=2, pady=2)
        Label(self.frame, text="Middle Initial: ").grid(row=1, sticky=W, ipady=2, pady=2)
        Label(self.frame, text="Last Name: ").grid(row=2, sticky=W, ipady=2, pady=2)
        Label(self.frame, text="Customer Type: ").grid(row=3, sticky=W, ipady=2, pady=2)
        Label(self.frame, text="Date (mm/dd/yyyy): ").grid(row=4, sticky=W, ipady=2, pady=2)

        self.fname_en = Entry(self.frame, width=30, textvariable=self.fname)
        self.mname_en = Entry(self.frame, width=30, textvariable=self.mname)
        self.lname_en = Entry(self.frame, width=30, textvariable=self.lname)
        self.payment_cb = Combobox(self.frame, textvariable=self.payment, width=27,
                                   values=("Drop In", "Punch Card", "Monthly", "Inactive"))
        self.payment_cb.set("Drop In")
        self.date_en = Entry(self.frame, width=30, textvariable=self.date)

        Frame(self.frame, width=5).grid(row=0,column=1,sticky=W)
        
        self.fname_en.grid(row=0,column=2,columnspan=2,sticky=W)
        self.mname_en.grid(row=1,column=2,columnspan=2,sticky=W)
        self.lname_en.grid(row=2,column=2,columnspan=2,sticky=W)
        self.payment_cb.grid(row=3,column=2,columnspan=2,sticky=W)
        self.date_en.grid(row=4,column=2,columnspan=2,sticky=W)
        
        ### buttons
        Button(self.frame, text='Cancel', width=10,
               command=self.wm_delete_window).grid(row=5, column=2, sticky=W, padx=10, pady=3)
        Button(self.frame, text='Submit', width=10,
               command=self.add_customer).grid(row=5, column=3, sticky=W)
        self.frame.pack(padx=10, pady=10)

        self.root.bind("<Return>", self.add_customer)
        self.fname_en.focus_set()

        if line: #preset values
            self.fname.set(line[1])
            self.mname.set(line[2])
            self.lname.set(line[0])
            self.payment_cb.set(line[3])
            self.date.set(line[4].strftime("%m/%d/%Y"))
        
        ### enable from base class
        self.enable()

    def add_customer(self, event=None):
        # validate and show errors
        if self.fname.get() == '':
            showerror("Error!", "First name field blank!")
        elif self.lname.get() == '':
            showerror("Error!", "Last name field blank!")
        elif self.mname.get() == '':
            showerror("Error!", "Middle initial field blank!")
        elif self.payment.get() not in ("Drop In", "Punch Card", "Monthly", "Inactive"):
            showerror("Error!", "Incorect Customer type!")
        elif not re.compile(r'[01]?\d/[0123]?\d/[12]\d{1,3}').search(self.date.get()):
            showerror("Error!", "Bad entry for date, use format mm/dd/yyyy")
        else:
            self.close = True
            
            # do work
            old, row = self.customers.find(str(self.lname.get()).strip(), str(self.fname.get()).strip(),
                                           str(self.mname.get()).strip())
            new = [str(self.lname.get()).strip(), str(self.fname.get()).strip(), str(self.mname.get()).strip(),
                   str(self.payment.get()).strip(), datetime.strptime(self.date.get(), "%m/%d/%Y")]
            
            if not old and self.close:
                self.new_customer_name = ' '.join([new[1],new[2],new[0]])
                self.customers.add(new)
            elif self.close and not self.edit:
                var = IntVar()
                
                diag = AlreadyExistsDialog(self.root, new, old, var)
                diag.show()
                if var.get() == 0: # edit
                    self.close = False
                if var.get() == 1: # replace
                    self.customers.replace(row, new)
                    self.new_customer_name = ' '.join([new[1],new[2],new[0]])
                elif var.get() == 2: # add duplicate
                    self.customers.add(new)
                    self.new_customer_name = ' '.join([new[1],new[2],new[0]])
            else:
                self.customers.replace(row, new)

            if self.close:
                self.root.quit()

class CustomerFrame(Frame):
    def __init__(self, master, customers, output_text, refresh):
        Frame.__init__(self, master)
        self.output_text = output_text
        self.refresh = refresh
        self.root = master
        self.customers = customers

        self.name = StringVar() #edit customer
        self.names = []
        self.ncd = NewCustomerDialog(self.root, self.customers, edit=True)

        self.fname = StringVar()
        self.lname = StringVar()
        self.mname = StringVar()
        self.payment = StringVar()
        self.date = StringVar()
        self.iconname="New Customer"

        lf = LabelFrame(self, text="New Customer")
        lf.grid(padx=5,pady=5,row=0,column=0,sticky='ew')
        
        ### dialog content        
        Label(lf, text="Name: ").grid(row=0,column=0,sticky='e',padx=(10,0),pady=(10,2))
        Label(lf, text="Type: ").grid(row=1,sticky='e',pady=2,padx=(10,0))
        Label(lf, text="Date: ").grid(row=1,column=2,sticky='e',ipady=2,padx=(10,0))

        self.fname_en = Entry(lf, width=20, textvariable=self.fname)
        self.mname_en = Entry(lf, width=4, textvariable=self.mname)
        self.lname_en = Entry(lf, width=20, textvariable=self.lname)
        self.payment_cb = Combobox(lf, textvariable=self.payment, width=12,
                                   values=("Drop In", "Punch Card", "Monthly", "Inactive"))
        self.date_en = Entry(lf, width=15, textvariable=self.date)
        
        self.fname_en.grid(row=0,column=1,sticky='ew',pady=(10,2))
        self.mname_en.grid(row=0,column=2,sticky='ew',pady=(10,2))
        self.lname_en.grid(row=0,column=3,sticky='ew',padx=(0,10),pady=(10,2))
        self.payment_cb.grid(row=1,column=1,sticky='ew')
        self.date_en.grid(row=1,column=3,columnspan=2,sticky='ew',padx=(0,10))
        
        ### buttons
        Button(lf, text='Reset Values', width=15,
               command=self.reset_values).grid(row=3,column=0,columnspan=2,sticky='ew',padx=10,pady=(2,10))
        Button(lf, text='Submit', width=15,
               command=self.add_customer).grid(row=3,column=3,sticky='ew',padx=(0,10),pady=(2,10))

        for i in range(4):
            lf.columnconfigure(i, weight=1)

        # edit customer
        lf = LabelFrame(self, text="Edit Customer")
        lf.grid(padx=5,pady=5,row=1,column=0,sticky='ew')

        Label(lf, text="Name: ").grid(row=0,column=0,sticky='e',pady=10,padx=(10,0))
        self.name_cb = Combobox(lf, textvariable=self.name, width=30, values=self.names)
        self.name_cb.grid(row=0,column=1,sticky='ew',pady=10)
        Button(lf, text="Edit",width=15,command=self.edit).grid(row=0,column=2,sticky='ew',padx=10,pady=10)

        for i in range(3):
            lf.columnconfigure(i,weight=1)

        self.columnconfigure(0,weight=1)

        self.fname_en.focus_set() #cursor goes here when frame is created
        self.update_names()
        self.reset_values() #zero out all values in new customer

    def edit(self):
        old_name = str(self.name.get())
        parsed = old_name.split(' ',2)
        (line,row) = self.customers.find(parsed[2],parsed[0],parsed[1])
        
        self.ncd.show(line)
        self.refresh() #refresh the global refresh
        name = ' '.join([self.ncd.fname.get(),self.ncd.mname.get(),self.ncd.lname.get()])
        self.output_text("+ - Modified: " + old_name + ' (' + line[3] + ') -> ' + name + " (" + self.ncd.payment.get() + ")\n")

    def update_names(self):
        self.populate_names()
        self.name_cb['values'] = self.names

    def populate_names(self):
        try:
            clist = self.customers.get_list()
        except IOError:
            self.output_text("! - " + self.customers.filename + " open in another application.\n")
            return
        clist.sort(key = lambda x: ', '.join(x[0:3]).lower())
        self.names = []
        for line in clist:
            self.names.append(' '.join([line[1],line[2],line[0]]))

    def reset_values(self):
        self.fname.set('')
        self.mname.set('')
        self.lname.set('')
        # blow out the field every time this is created
        self.date.set(date.today().strftime("%m/%d/%Y"))
        self.payment_cb.set("Drop In")

    def add_customer(self, event=None):
        # validate and show errors
        if self.fname.get() == '':
            showerror("Error!", "First name field blank!")
        elif self.lname.get() == '':
            showerror("Error!", "Last name field blank!")
        elif self.mname.get() == '':
            showerror("Error!", "Middle initial field blank!")
        elif self.payment.get() not in ("Drop In", "Punch Card", "Monthly", "Inactive"):
            showerror("Error!", "Incorect Customer type!")
        elif not re.compile(r'[01]?\d/[0123]?\d/[12]\d{1,3}').search(self.date.get()):
            showerror("Error!", "Bad entry for date, use format mm/dd/yyyy")
        else:
            # do work
            name = ' '.join([self.fname.get(),self.mname.get(),self.lname.get()])
            old, row = self.customers.find(str(self.lname.get()).strip(), str(self.fname.get()).strip(),
                                           str(self.mname.get()).strip())
            new = [str(self.lname.get()).strip(), str(self.fname.get()).strip(), str(self.mname.get()).strip(),
                   str(self.payment.get()).strip(), datetime.strptime(self.date.get(), "%m/%d/%Y")]
            
            if not old: #add customer
                self.customers.add(new)
                self.output_text("+ - New Customer: " + name + " (" + self.payment.get() + ")\n")
                self.refresh()
            else:
                var = IntVar()
                
                diag = AlreadyExistsDialog(self.root, new, old, var)
                diag.show()
                if var.get() == 0: # edit
                    pass
                if var.get() == 1: # replace customer
                    self.customers.replace(row, new)
                    self.output_text("+ - Modified: " + name + " (" + self.payment.get() + ")\n")
                    self.refresh()

class AlreadyExistsDialog(Dialog):
    def __init__(self, master_frame, new, old, variable, class_=None, relx=0.5, rely=0.3):
        Dialog.__init__(self, master_frame, 'Warning!',
                        class_, relx, rely)
        self.new = new
        self.old = old
        self.variable = variable
        
        self.text = "Warning: Customer Found with the same name.\n" + \
        "\nNew record: " + new[1] + " " + new[2] + " " + new[0] + \
        " (" + new[3] + ") - " + new[4].strftime("%m/%d/%Y") + \
        "\nOld record: " + old[1] + " " + old[2] + " " + old[0] + \
        " (" + old[3] + ") - " + old[4].strftime("%m/%d/%Y") + \
        "\n\n Cancel to edit entry, Override to replace old entry."
        

    def show(self):
        self.setup()

        self.msg = Message(self.root, text=self.text, aspect=400)
        self.msg.pack(expand=True, fill=BOTH)

        self.frame = Frame(self.root)
        
        b1 = Button(self.frame, text="Cancel", command=self.cancel)
        b1.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        b2 = Button(self.frame, text="Override", command=self.replace)
        b2.pack(side=LEFT, fill=BOTH, expand=True, padx=10)
        #b3 = Button(self.frame, text="Add Duplicate", command=self.add_duplicate)
        #b3.pack(side=LEFT, fill=BOTH, expand=True)
        
        self.frame.pack(padx=10, pady=10)

        self.root.bind('<Return>', self.cancel)
        
        self.enable()

    def cancel(self, event=None):
        self.variable.set(0)
        self.root.quit()

    def replace(self):
        self.variable.set(1)
        self.root.quit()

    def add_duplicate(self):
        self.variable.set(2)
        self.wm_delete_window()
        

class Customers:
    def __init__(self, filename='jim_info.xlsx', sheet_name='Customers'):
        self.wb = load_workbook(filename)
        self.filename = filename
        self.sh = self.wb.get_sheet_by_name(sheet_name)
        self.sheet_name = sheet_name
        if not self.sh:
            print "Error opening " + sh_name + " sheet."

    def reload_file(self):
        self.wb = load_workbook(self.filename)
        self.sh = self.wb.get_sheet_by_name(self.sheet_name)

    def add(self, new):
        ''' new = [lname, fname, email, ptype, date_created] '''
        self.sh.garbage_collect()
        # Append new customer
        self.sh.append(new)
        # set the date style
        self.sh.cell(row=(self.sh.get_highest_row()-1),
                     column=4).style.number_format.format_code = 'm/d/yyyy'    
        
        # save the file
        self.wb.save(self.filename)

    def replace(self, row, new):
        ''' new = [lname, fname, mname, ptype, date_created] '''
        for col in range(5):
            self.sh.cell(row=row,column=col).value = new[col]

        self.sh.cell(row=row,
                     column=4).style.number_format.format_code = 'm/d/yyyy'  
        # save the file
        self.wb.save(self.filename)

    def get_list(self):
        self.reload_file()
        output = []
        for r in range(1, self.sh.get_highest_row()):
            points = self.sh.range("A1:E1",row=r)
            values = []
            for i in range(5):
                values.append(str(points[0][i].value).strip())
            
            output.append(values)
        return output

    def get_dict(self):
        values = self.get_list()
        output = {}
        for r in range(len(values)):
            name = ', '.join(values[r][:2]).strip()
            output[name] = values[r][2:]
        return output

    def get_type(self, name):
        parsed = name.split(' ',2)
        (line,row) = self.find(parsed[2],parsed[0],parsed[1])
        if not line: return "Customer Not Found"
        return line[3]
    
    def find(self, lname, fname, mname):
        self.reload_file()
        output = None
        row = None
        for r in range(self.sh.get_highest_row()):
            if (lname == str(self.sh.cell(row=r, column=0).value).strip() and
                fname == str(self.sh.cell(row=r, column=1).value).strip() and 
                mname == str(self.sh.cell(row=r, column=2).value).strip()):
                temp = self.sh.range("A1:E1",row=r)
                output = []
                for i in range(5):
                    if i != 4:
                        output.append(str(temp[0][i].value).strip())
                    else:
                        output.append(temp[0][i].value)
                row = r
                break
        return output, row

class EditCustomerDialog(NewCustomerDialog):
    def __init__(self, master, customers, output_text, class_=None, relx=0.5, rely=0.3):
        Dialog.__init__(self, master, 'Edit Customer',
                        class_, relx, rely)
        self.customers = customers
        self.output_text = output_text
        self.fname = StringVar()
        self.lname = StringVar()
        self.mname = StringVar()
        self.payment = StringVar()
        self.date = StringVar()

        already_exists_show = False

        self.new_customer_name = None

    def show(self, lname, fname, mname, payment, date):
        self.fname.set(fname)
        self.mname.set(mname)
        self.lname.set(lname)
        self.payment.set(payment)
        self.date.set(date)

        NewCustomerDialog.show(self)

    def add_customer(self, event=None):
        NewCustomerDialog.add_customer(self,event)
        if self.close:
            self.output_text("+ - Modified: " + name + " (" + self.payment.get() + ")\n")

def output_text(text):
    print text,

def refresh():
    print "refreshing...."

if __name__ == '__main__':
    c = Customers()
    pprint(c.get_list())
    
    root = Tk()
    ncf = CustomerFrame(root, c, output_text, refresh)
    ncf.grid(sticky='nsew')
    root.rowconfigure(0,weight=1)
    root.columnconfigure(0,weight=1)

    root.mainloop()

    # c = Customers()
    # pprint(c.get_list())
    
    # root = Frame()
    # ncd = NewCustomerDialog(root, c)
    # Button(root,text='New Customer',command=ncd.show).pack()
    # root.pack()
    # root.mainloop()
