'''
scripts for generating reports for jim tracker
'''

from openpyxl import Workbook, load_workbook
from openpyxl.style import Border, Fill
from openpyxl.cell import get_column_letter
from pprint import pprint
from calendar import Calendar
from time import strftime
from datetime import time,datetime
from os import chdir, listdir, getcwd, system
import re,sys

from ttk import Frame,Label,Combobox,Label,Button,LabelFrame
from Tkinter import StringVar,Tk
import tkFileDialog

from customer import Customers
from payment import Payments

class ReportsFrame(Frame):
    def __init__(self,master,customers,payments,output_text):
        Frame.__init__(self,master)
        self.customers = customers
        self.payments = payments
        self.master = master
        self.output_text = output_text

        self.year_months = find_years_months(getcwd()) # use cwd, this should be set

        self.year = StringVar()
        self.month = StringVar()
        self.years = sorted(self.year_months.keys())
        self.months = []

        lf = LabelFrame(self, text="Generate Report")
        lf.grid(padx=5,pady=5,row=0,column=0,sticky='ew')

        Label(lf,text="Year: ").grid(row=0,column=0,sticky='e',padx=(10,0),pady=(10,2))
        Label(lf,text="Month: ").grid(row=1,column=0,sticky='e',padx=(10,0),pady=2)

        self.year_cb = Combobox(lf,textvariable=self.year,width=12,values=self.years,state='readonly')
        self.month_cb = Combobox(lf,textvariable=self.month,width=12,values=self.months,state='readonly')
        
        self.year_cb.grid(row=0,column=1,sticky='w',padx=(0,10),pady=(10,2))
        self.month_cb.grid(row=1,column=1,sticky='w',padx=(0,10),pady=2)

        btn = Button(lf,text="Save Report",command=self.report,width=30)
        btn.grid(row=2,column=0,columnspan=2,sticky='n',pady=(2,10),padx=10)

        #configure the grid to expand
        self.columnconfigure(0,weight=1)
        lf.rowconfigure(0,weight=1)
        lf.rowconfigure(1,weight=1)
        lf.columnconfigure(0,weight=1)
        lf.columnconfigure(1,weight=1)

        self.year_cb.bind('<<ComboboxSelected>>',self.year_selected)

        self.update() #update the values

    def report(self):
        '''
        generate the report, run by clicking Button
        '''
        if self.year.get() is '':
            self.output_text("! - Please Select a Year")
            return
        if self.month.get() is '':
            self.output_text("! - Please select a Month")
            return

        year = self.year.get()
        inputf = 'jim_data' + year + '.xlsx'
        month = self.month.get()
        outputf_def = month + year + '_report.xlsx'
        outputf = tkFileDialog.asksaveasfilename(parent=self,
            defaultextension='.xlsx', initialfile=outputf_def)
        if outputf is '': return #file not selected

        #output report
        month_report(inputf,month,year,outputf,self.customers,self.payments)

        self.output_text("* - " + self.month.get() + ' ' + self.year.get() + ' report saved to: ' + outputf + '\n')

        if sys.platform is 'debian': 
            system('open ' + outputf)
        else:
            system(outputf) # open the file

    def update(self):
        '''
        method for updating values when things change
        '''
        self.year_months = find_years_months(getcwd()) # use cwd, this should be set
        self.years = sorted(self.year_months.keys())
        self.months = ['']
        if len(self.years) == 0: self.years = ['']
        self.year_cb['values'] = self.years
        self.month_cb['values'] = self.months
        self.month_cb.current(0)

    def year_selected(self,event=None):
        '''
        run when year is year is selected 
        '''
        self.months = self.year_months[self.year.get()]
        self.month_cb['values'] = self.months
        self.month_cb.current(0)

def workouts_this_month(customer,log_file,month=strftime("%B")):
    '''
    laod the log file and count number of workouts done by customer
    '''
    wb = load_workbook(log_file)
    sh = wb.get_sheet_by_name(month)

    data = get_data(sh)
    count = len( [row for row in data if row[3] == customer] )
    return count

def get_data(sh):
    '''
    converts sheet into array of arrays (rows of columns)
    '''
    if not sh:
        return None
    #get the data
    data = [[cell.value for cell in row] for row in sh.rows]
    return data

def find_years_months(root):
    """ finds all the data files in root directory 
    returns dictionary of years maped to lists of months

    throws IOError for any file open that matches the pattern
    """
    files = listdir(root)
    # print files

    months = ['January', 'February', 'March', 'April', 'May', 'June', 
    'July', 'August', 'September', 'October', 'November', 'December']

    output = dict()

    data_re = re.compile(r"jim_data\d\d\d\d[.]xlsx")
    for filename in files:
        match = data_re.match(filename)
        if match:
            wb = load_workbook(match.group())
            sheet_names = wb.get_sheet_names()
            ordered_months = []
            for m in months:
                if m in sheet_names:
                    ordered_months.append(m)
            output[match.group()[8:12]] = ordered_months

    return output

def generate_info_file():
    """ generates the jim_info.xlsx file in working directory """
    class_labels = ['Mon', 'type', 'Tue', 'type', 'Wed', 'type', 'Thurs', 'type', 'Fri', 'type', 'Sat', 'type', 'Sun', 'type']
    wb = Workbook()
    sh = wb.get_active_sheet()
    sh.title = "Schedule"
    sh.append(class_labels)
    label_format(sh,len(class_labels))
    for col in range(1,len(class_labels),2):
        for row in range(0,14):
            cell = sh.cell(row=row,column=col).style
            cell.borders.right.border_style = Border.BORDER_THIN
    for col in range(0,len(class_labels)):
        cell = sh.cell(row=14,column=col).style
        cell.borders.top.border_style = Border.BORDER_THIN

    sh.cell(row=1,column=0).value = time(10,0)
    sh.cell(row=1,column=0).style.number_format.format_code = 'h:mm'
    sh.cell(row=1,column=1).value = "Caveman"

    customers_labels = ['Last', 'First', 'Middle', 'Type', 'Date', 'Joined',]
    sh = wb.create_sheet(title="Customers")
    sh.append(customers_labels)
    label_format(sh,len(customers_labels))

    sh = wb.create_sheet(title="Admin")
    sh.append(['This sheet is used for internal variables, do not modify or remove!'])
    sh.append(["Last Opened", datetime.today()])

    wb.save("jim_info.xlsx")

def month_report(log_file,month,year,output_file,customers,payments):
    #read the log
    wb = load_workbook(log_file)
    sh = wb.get_sheet_by_name(month)

    if not sh:
        # error - sheet doesn't exist
        return False

    #get the data
    data = get_data(sh)

    wb = Workbook()
    sh = wb.get_active_sheet()
    sh.title = "Customers"
    customers_report(data,sh,customers,payments,month,year)

    sh = wb.create_sheet(title='Classes')
    class_sheet = class_report(data,sh)
    
    #write new workbook
    wb.save(output_file)

def label_format(sh,columns,row=0,border='bottom'):
    #format top row
    for col in range(columns):
        cell = sh.cell(row=row, column=col).style
        cell.fill.fill_type = Fill.FILL_SOLID
        cell.fill.start_color.index = "FFDDD9C4"
        if border == 'bottom':
            cell.borders.bottom.border_style = Border.BORDER_THIN
        if border == 'top':
            cell.borders.top.border_style = Border.BORDER_THIN

def customers_report(data,sh,c,p,month,year):
    '''
    Build the Customers, # of workouts report
    data, log data 
    sh, output sheet
    c, customer class  
    p, payments class
    '''

    #create a dictionary of customers this month
    customers = dict.fromkeys(set([str(x[3]) for x in data[1:]]))
    for key in customers:
        customers[key] = [row for row in data if row[3] == key]

    report_data = [('Customer','# of Workouts','Type','Note')] + \
                  sorted([(key,str(len(value)),c.get_type(key),customer_note(key,c,p,month,year)) for (key,value) in customers.items()],
                         key=lambda pair: int(pair[1]), reverse=True)

    #write data
    for values in report_data:
        sh.append(values)
        color = None
        if values[3] is "Unpaid":
            color = "00ff0000"
        elif values[3] is "Paid":
            color = "0000ff00"

        if color:
            for col in range(4):
                cell = sh.cell(row=sh.get_highest_row()-1,column=col).style
                cell.fill.fill_type = Fill.FILL_SOLID
                cell.fill.start_color.index = color

    #format labels
    label_format(sh,4)
    # set column width

    for i, column_width in enumerate([25, 13, 18, 25]):
        sh.column_dimensions[get_column_letter(i+1)].width = column_width

def customer_note(name,customers,payments,month,year):
    '''
    name - customer name
    customers - Customers object
    payments - Payments object
    '''
    ctype = customers.get_type(name) # the customer type
    if ctype == "Customer Not Found":
        return "Add customer to info file."

    if ctype == "Monthly":
        if payments.has_paid_monthly(name,month,year):
            return "Paid"
        else:
            return "Unpaid"

    if ctype == "Punch Card":
        return "Remaining Punches: " + str(payments.get_remaining_punches(name,year,month))

    return ""


def class_report(data,sh):
    '''
    creates a calendar that shows the number of customers at each workout
    '''
    dates = map(lambda x: x.date(), sorted(list(set([x[0] for x in data[1:]]))))
    report_data = dict.fromkeys(dates)
    for day in report_data:
        report_data[day] = dict.fromkeys(set([(x[1],x[2]) for x in data[1:] if x[0].date()==day]))
        for workout in report_data[day]:
            #set is used to prevent duplicates from being counted
            report_data[day][workout] = len(set([x[3] for x in data[1:] if (x[0].date()==day and (x[1],x[2])==workout)]))
    weekdays = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    line = []
    for wkd in weekdays:
        line += [wkd,'Class','# Att']
    sh.append(line)
    label_format(sh,len(line))
    first = True
    line = []
    workouts = {}
    for day in Calendar().itermonthdates(dates[0].year,dates[0].month):
        if day.weekday() == 0 and not first:
            sh.append(line)
            label_format(sh,len(line),sh.get_highest_row()-1,'top')
            if workouts:
                row = sh.get_highest_row()
                for wkd in workouts:
                    for i,workout in enumerate(workouts[wkd]):
                        sh.cell(row=row+i,column=wkd*3).value = workout[0].strftime("%H:%M")
                        sh.cell(row=row+i,column=wkd*3+1).value = workout[1]
                        sh.cell(row=row+i,column=wkd*3+2).value = workout[2]
            line = []
            workouts = {}
            
        if day in dates:
            line.append(day.day)
            line.append('')
            line.append('')
            date_data = [day.day,'','']
            for (time,workout),num in report_data[day].items():
                if workouts.has_key(day.weekday()):
                    workouts[day.weekday()].append([time,workout,num])
                else:
                    workouts[day.weekday()] = [[time,workout,num]]
        else:
            line.append(day.day)
            line.append('')
            line.append('')

        first = False
    sh.append(line)
    label_format(sh,len(line),sh.get_highest_row()-1,'top')
    if workouts:
        row = sh.get_highest_row()
        for wkd in workouts:
            for i,workout in enumerate(workouts[wkd]):
                sh.cell(row=row+i,column=wkd*3).value = workout[0].strftime("%H:%M")
                sh.cell(row=row+i,column=wkd*3+1).value = workout[1]
                sh.cell(row=row+i,column=wkd*3+2).value = workout[2]

    for col in range(0,sh.get_highest_column()+1,3):
        for row in range(sh.get_highest_row()):
            cell = sh.cell(row=row, column=col).style
            cell.borders.left.border_style = Border.BORDER_THIN

    auto_column_width(sh)
                

def auto_column_width(worksheet):
    raw_data = worksheet.range(worksheet.calculate_dimension())
    data = [[str(x.value) for x in row] for row in raw_data]
    column_widths = []
    for row in data:
        for i, cell in enumerate(row):
            if len(column_widths) > i:
                if len(cell) > column_widths[i]:
                    column_widths[i] = len(cell)
            else:
                column_widths += [len(cell)]

    for i, column_width in enumerate(column_widths):
        worksheet.column_dimensions[get_column_letter(i+1)].width = column_width

def output_text(text):
    print text,

def refresh():
    print "refreshing...."

def test1():
    inputf = 'jim_data2013.xlsx'
    month = 'August'
    year = '2013'
    c = Customers()
    p = Payments()
    outputf = month + '_report.xlsx'
    month_report(inputf,month,year,outputf,c,p)
    # print workouts_this_month("Dave L Sanders", inputf, month) 

    # #test generate info file
    # chdir("C:\\Users\\tyler.weaver\\Projects")
    # generate_info_file()

    # test find data files
    # print find_years_months(getcwd())

def test_frame():
    c = Customers()
    p = Payments()

    root = Tk()
    rf = ReportsFrame(root,c,p,output_text)
    rf.grid(sticky='nsew')
    root.rowconfigure(0,weight=1)
    root.columnconfigure(0,weight=1)

    root.mainloop()

if __name__ == '__main__':
    test1()
    # test_frame()