'''
Payment objects for jim tracker

jim_payments<year>.xlsx (book)
    <month> (sheet)
        Monthly (column 0,1)
            Customer - 0
            Date - 1
        Punch Card (column 2,3,4)
            Customer - 2
            Date - 3
            Remaining - 4
        Drop In (column 5,6)
            Customer - 5
            Date - 6
'''
#standard python
from time import strftime
from datetime import datetime
#openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.shared.exc import InvalidFileException
from openpyxl.style import Border, Fill
from openpyxl.cell import get_column_letter
#tkinter
from ttk import Frame,Label,Entry,Combobox,LabelFrame,Button
from Tkinter import StringVar,Tk
from tkMessageBox import showerror
#jim tracker
from customer import Customers

class PaymentFrame(Frame):
    """docstring for PaymentFrame"""
    def __init__(self, master, customers, payments, output_text, refresh):
        Frame.__init__(self, master)

        self.refresh = refresh
        self.master = master
        self.output_text = output_text
        self.customers = customers
        self.payments = payments

        self.pname = StringVar()
        self.pnames = []
        self.mname = StringVar()
        self.mnames = []
        self.date = StringVar()
        self.nmonths = StringVar()
        self.punches = StringVar()

        self.nmonths.set('1')
        self.punches.set(str(10))
        self.date.set(strftime("%m/%d/%Y"))

        self.columnconfigure(0,weight=1)

        # Monthly Customers
        monthly_lf = LabelFrame(self, text="Monthly Customers Payment")
        monthly_lf.grid(padx=5,pady=5,row=0,column=0,sticky='ew')
        
        Label(monthly_lf,text="Name:").grid(row=0,column=0,sticky='e',padx=(10,0),pady=(10,2))
        Label(monthly_lf,text="Date:").grid(row=0,column=3,sticky='e',padx=(10,0),pady=(10,2))
        Label(monthly_lf,text="# Months:").grid(row=1,column=0,columnspan=2,sticky='e',padx=(10,0),pady=(2,10))
        self.mname_cb = Combobox(monthly_lf,textvariable=self.mname,width=20,values=self.mnames,
            state='readonly')
        self.mname_cb.grid(row=0,column=1,columnspan=2,sticky='ew',pady=(10,2))
        Entry(monthly_lf,textvariable=self.date,width=15).grid(row=0,column=4,sticky='ew',padx=(0,10),pady=(10,2))
        Entry(monthly_lf,textvariable=self.nmonths).grid(row=1,column=2,sticky='ew',pady=(2,10))
        Button(monthly_lf,text='Submit',command=self.monthly_payment).grid(row=1,column=4,sticky='ew',padx=(0,10),pady=(2,10))

        for i in range(5):
            monthly_lf.columnconfigure(i,weight=1)

        # Punch Card Customers
        puch_lf = LabelFrame(self, text="Punch Card Customers (Purchace Card)")
        puch_lf.grid(padx=5,pady=5,row=1,column=0,sticky='ew')

        Label(puch_lf,text="Name:").grid(row=0,column=0,sticky='e',padx=(10,0),pady=(10,2))
        Label(puch_lf,text="Punches:").grid(row=0,column=2,sticky='e',pady=(10,2))
        self.pname_cb = Combobox(puch_lf,textvariable=self.pname,width=20,values=self.pnames,state='readonly')
        self.pname_cb.grid(row=0,column=1,sticky='ew',pady=(10,2))
        Entry(puch_lf,textvariable=self.punches,width=15).grid(row=0,column=3,sticky='ew',padx=(0,10),pady=(10,2))
        Button(puch_lf,text='Submit',command=self.new_punchcard).grid(row=3,column=3,sticky='ew',padx=(0,10),pady=(2,10))

        for i in range(4):
            puch_lf.columnconfigure(i,weight=1)

        self.update_names()

    def monthly_payment(self):
        nextpayment_date = datetime.strptime(self.date.get(), "%m/%d/%Y")
        try:
            self.payments.monthly_payment(self.mname.get(),datetime.strptime(self.date.get(), "%m/%d/%Y"), int(self.nmonths.get()))
            for i in range(int(self.nmonths.get())):
                nextpayment_date = increment_month(nextpayment_date)
        except IOError:
            showerror("Error writting to file", "Please close " + self.payments.filename + " and press OK.")
        except ValueError:
            self.output_text("! - Bad value for date (mm/dd/yyyy) or # Months \n")
        else:
            if int(self.nmonths.get()) > 1:
                self.output_text("$ - " + self.mname.get() + " Paid for " + self.nmonths.get() + " months\n")
            else:
                self.output_text("$ - " + self.mname.get() + " Paid for 1 month\n")

            self.output_text("$ - Next Payment Due: " + nextpayment_date.strftime("%B %d, %Y") + "\n")
            self.refresh()

    def new_punchcard(self):
        try:
            punches = int(self.punches.get())
            if punches < 1 or punches > 10:
                raise ValueError
            self.payments.new_punchcard(self.pname.get(), punches = punches)
            self.output_text("$ - New Puncard: " + self.pname.get() + "\n")
        except IOError:
            showerror("Error writting to file", "Please close " + self.payments.filename + " and press OK.")
        except ValueError:
            self.output_text("! - Bad value for number of punches: " + self.punches.get() + "\n")
        else:
            self.reset_punchcard()
            self.refresh()

    def reset_punchcard(self):
        self.punches.set(str(10))

    def reset_monthly(self):
        self.date.set(strftime("%m/%d/%Y"))
        self.nmonths.set('1')

    def update_names(self):
        '''
        run by refresh
        '''
        self.populate_names()
        self.mname_cb['values'] = self.mnames
        if len(self.mnames) > 0: self.mname_cb.current(0)
        self.pname_cb['values'] = self.pnames
        if len(self.pnames) > 0: self.pname_cb.current(0)
        self.reset_punchcard()
        self.reset_monthly()
        
    def populate_names(self):
        # try:
        clist = self.customers.get_list()
        # except IOError:
        #     self.output_text("! - " + self.customers.filename + " open in another application.\n")
        #     return
        clist.sort(key = lambda x: ', '.join(x[0:3]).lower())
        self.mnames = [' '.join([line[1],line[2],line[0]]) for line in clist if line[3]=='Monthly']
        self.pnames = [' '.join([line[1],line[2],line[0]]) for line in clist if line[3]=='Punch Card']


#jim_tracker
class Payments:
    def __init__(self):
        '''
        Constructor for Payments.
        Throws IOError if file is open in another application.
        '''
        self.refresh()

    def refresh(self):
        '''
        reset all variables
        '''
        self.month = strftime("%B")
        self.year = strftime("%Y")
        self.filename = 'jim_payments' + str(self.year) + '.xlsx'

        self.wb = self.open_workbook()
        self.sh = self.open_sheet()
        
    def open_workbook(self,year=None):
        '''
        helper method for initalization
        Throws IOError if file is open in another application.

        remember to refresh after save if new file!
        '''
        if not year: 
            year = self.year
            filename = self.filename
        else:
            self.filename = 'jim_payments' + str(year) + '.xlsx'

        try:
            workbook = load_workbook(self.filename)
        except InvalidFileException: #create new file!
            workbook = Workbook()
            sh = workbook.get_sheet_by_name("Sheet")
            workbook.remove_sheet(sh)

        return workbook

    def open_sheet(self,workbook=None,month=None):
        '''
        helper method for initalization
        '''
        default_month = False
        if not month: 
            month = self.month
            default_month = True
        if not workbook: workbook = self.wb
        sheet = workbook.get_sheet_by_name(month)
        if not sheet: # new month!
            workbook.create_sheet(title=month)
            sheet = workbook.get_sheet_by_name(month)
            sheet.append(['Monthly', '', 'Punch Card', '', '', 'Drop In', ''])
            sheet.append(['Customer', 'Date', 'Customer', 'Date', 'Remaining', 'Customer', 'Date'])
            for row in range(2):
                for col in range(7):
                    cell = sheet.cell(row=row, column=col).style
                    cell.fill.fill_type = Fill.FILL_SOLID
                    cell.fill.start_color.index = "FFDDD9C4"
                    cell.borders.bottom.border_style = Border.BORDER_THIN

            # set column widths
            column_widths = [20,12,20,12,12,20,12]
            for i, column_width in enumerate(column_widths):
                sheet.column_dimensions[get_column_letter(i+1)].width = column_width
            
            if default_month:
                # auto card import
                today = datetime.today()
                old = None
                if today.month == 1: #january
                    old = datetime(today.year-1, 12, 1)
                else: # other months
                    old = datetime(today.year, today.month-1, 1)

                self.update_cards(old.strftime("%B"), old.strftime("%Y"))

        return sheet

        self.format_save()

    def monthly_payment(self, customer, date = datetime.today(), nmonths = 1):
        '''
        enters a new monthly Payment
        '''
        if date.month != datetime.today().month:
            # we only care if the month is different
            if date.year != datetime.today().year:
                #different year, open different workbook
                workbook = self.open_workbook(date.strftime("%Y"))
                sheet = self.open_sheet(workbook, date.strftime("%B"))
            else:
                #same year, just open current month
                workbook = self.wb
                sheet = self.open_sheet(month=date.strftime("%B"))
        else:
            sheet = self.sh
            workbook = self.wb

        #find the next empty line (row value)
        row = 2
        cust_column = 0
        while sheet.cell(row=row,column=cust_column).value != None:
            row += 1

        sheet.cell(row=row,column=0).value = customer
        sheet.cell(row=row,column=1).value = date

        self.format_save(workbook,sheet)

        if nmonths > 1:
            self.monthly_payment(customer,increment_month(date),nmonths-1)

    def new_punchcard(self, customer, date = None, punches = 10):
        '''
        creates a new punch card entry
        '''
        #logic for custom date entries
        workbook = self.wb
        sheet = self.sh
        if date == None:
            date = datetime.today()
        else:
            month = date.strftime("%B")
            year = date.strftime("%Y")
            if year != self.year:
                workbook = self.open_workbook(year)
                sheet = self.open_sheet(workbook,month)
            elif month != self.month:
                sheet = self.open_sheet(self.wb,month)
            else:
                sheet = self.sh

        #find the next empty line (row value)
        row = 2
        cust_column = 2
        while sheet.cell(row=row,column=cust_column).value != None:
            row += 1

        sheet.cell(row=row,column=2).value = customer
        sheet.cell(row=row,column=3).value = date
        sheet.cell(row=row,column=4).value = punches

        self.format_save(workbook,sheet)

    def punch(self, customer):
        '''
        punches a punchcard 
        1. finds punch card with remaining punches
        2. punches
        3. returns remaining punchs

        returns None if card does not exist or no remaining punches
        '''
        #find the punchcard (customer) with remaining punches
        row = 2
        cust_column = 2
        punch_column = 4
        found_zero = False
        for row in range(self.sh.get_highest_row()+1):
            if self.sh.cell(row=row,column=cust_column).value == customer:
                if int(self.sh.cell(row=row,column=punch_column).value) > 0:
                    break
                if int(self.sh.cell(row=row,column=punch_column).value) == 0:
                    found_zero = True
            if self.sh.cell(row=row,column=cust_column).value == None: # not found
                if found_zero: return 0
                else: return None
            
        punch = int(self.sh.cell(row=row,column=punch_column).value) - 1
        self.sh.cell(row=row,column=punch_column).value = punch

        self.format_save()

        return punch

    def get_remaining_punches(self, customer, year=strftime("%Y"), month=strftime("%B")):
        #find the punchcard (customer) with remaining punches
        workbook = self.open_workbook(year)
        sheet = self.open_sheet(workbook, month)
        row = 2
        cust_column = 2
        punch_column = 4
        punches = 0
        for row in range(sheet.get_highest_row()+1):
            if sheet.cell(row=row,column=cust_column).value == customer:
                punches += int(sheet.cell(row=row,column=punch_column).value)

        return punches

    def drop_in(self, customer, date = datetime.today()):
        '''
        enters a drop in Payment
        '''
        #find the next empty line (row value)
        row = 2
        cust_column = 5
        while self.sh.cell(row=row,column=cust_column).value != None:
            row += 1

        self.sh.cell(row=row,column=5).value = customer
        self.sh.cell(row=row,column=6).value = date

        self.format_save()

    def update_cards(self, from_month, year = None):
        '''
        copies in old punches from previous month
        can copy in from old years (different filename)

        returns None if month or year does not exist
        returns number of cards copied otherwise
        '''
        if not year:
            year = self.year

        cards = []

        if year != self.year: #different year
            from_filename = 'jim_payments' + str(year) + '.xlsx'
            try:
                from_wb = load_workbook(from_filename)
            except InvalidFileException: #file does not exist
                return None
        else:
            from_wb = self.wb

        from_sh = from_wb.get_sheet_by_name(from_month)
        if not from_sh: # month does not exist
            return None

        #get the data
        row = 2
        cust_column = 2
        date_column = 3
        punch_column = 4
        while from_sh.cell(row=row,column=cust_column).value != None:
            if from_sh.cell(row=row,column=punch_column).value != 0:
                cards.append([from_sh.cell(row=row,column=cust_column).value,
                    from_sh.cell(row=row,column=date_column).value,
                    from_sh.cell(row=row,column=punch_column).value])
            row += 1

        # add the values to current sheet (tests if they exist first in current sheet (date validation))
        current_cards = {}
        row = 2
        while self.sh.cell(row=row,column=cust_column).value != None:
            current_cards[self.sh.cell(row=row,column=cust_column).value] = self.sh.cell(row=row,column=date_column).value
            row += 1
        
        cards_copied = 0
        for card in cards:
            if card[0] in current_cards:
                if current_cards[card[0]] == card[1]: #same punch card
                    continue
            self.new_punchcard(card[0],datetime.today(),card[2])
            cards_copied += 1

        return cards_copied

    def has_paid_monthly(self, customer, month, year):
        '''
        Returns true if customer has paid monthly dues
        '''
        workbook = self.open_workbook(year)
        sheet = self.open_sheet(workbook, month)

        row = 2
        cust_column = 0
        while sheet.cell(row=row,column=cust_column).value != None:
            if customer == sheet.cell(row=row,column=cust_column).value:
                return True
            row += 1

        return False

    def format_save(self, workbook=None, sheet=None):
        if not workbook: workbook = self.wb
        if not sheet: sheet = self.sh
        sheet.garbage_collect()
        self.formating(sheet)
        workbook.save(self.filename)
        # self.refresh()

    def formating(self,sheet):
        '''
        apply formatting to new changes
        '''
        columns = [1, 4, 6]
        for row in range(0, sheet.get_highest_row()):
            for col in columns:
                cell = sheet.cell(row=row, column=col).style
                cell.borders.right.border_style = Border.BORDER_THIN
            sheet.cell(row=row,column=1).style.number_format.format_code = 'm/d/yyyy'
            sheet.cell(row=row,column=3).style.number_format.format_code = 'm/d/yyyy'
            sheet.cell(row=row,column=4).style.number_format.format_code = '0'
            sheet.cell(row=row,column=6).style.number_format.format_code = 'm/d/yyyy'


def test1():
    p = Payments()

    monthly_customers = ['Tyler J Weaver', 'Marcus T Weaver']
    p.monthly_payment(monthly_customers[0])
    p.monthly_payment(monthly_customers[1])

    # for c in monthly_customers:
    #     print c
    #     if p.has_paid_monthly(c): 
    #         print "Paid"
    #     else:
    #         print "Unpaid"

    # punch_card_customers = ['Brad Bradley', 'Sam P Frank']

    # for c in punch_card_customers:
    #     p.new_punchcard(c)

    # for x in range(4):
    #     print punch_card_customers[0], p.punch(punch_card_customers[0])

    # for x in range(7):
    #     print punch_card_customers[1], p.punch(punch_card_customers[1])

    # drop_in_customers = ['Tom R Jones', 'Rodgers P Smith']

    # p.drop_in(drop_in_customers[0])
    # p.drop_in(drop_in_customers[1])

def increment_month(input_date):
    '''
    increments by one month and sets the day value to the 5th
    '''
    output = None
    if input_date.month < 12:
        output = datetime(input_date.year,input_date.month+1,5)
    else:
        output = datetime(input_date.year+1,1,5)
    return output

def output_text(text):
    print text

def refresh():
    print "refresh..."

def test2():
    ''' monthly payment test '''
    p = Payments()

    monthly_customers = ['Tyler J Weaver', 'Marcus T Weaver']
    p.monthly_payment(monthly_customers[0])
    p.monthly_payment(monthly_customers[1])
    p.monthly_payment(monthly_customers[0],datetime(2013,9,3))
    p.monthly_payment(monthly_customers[1],datetime(2014,2,2))
    p.monthly_payment(monthly_customers[0],datetime(2012,12,4))
    p.monthly_payment(monthly_customers[1],datetime(2013,7,4))

def test3():
    ''' test updating punch cards 
    not complete and working
    '''
    p = Payments()

    punch_card_customers = ['Brad Bradley', 'Sam P Frank']

    for c in punch_card_customers:
        p.new_punchcard(c,date=datetime(2013,7,1))

    p.update_cards("July")

def output_text(text):
    print text,

def refresh():
    print "refreshing...."

def testFrame():
    c = Customers()
    p = Payments()

    root = Tk()

    pf = PaymentFrame(root, c, p, output_text, refresh)
    pf.grid(sticky='nsew')
    root.rowconfigure(0,weight=1)
    root.columnconfigure(0,weight=1)

    root.mainloop()

if __name__ == '__main__':
    testFrame()