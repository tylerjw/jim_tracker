'''
jim tracker main application

'''

#tkinter
from ttk import Notebook,Frame,Label
from Tkinter import Text,Menu
from ScrolledText import ScrolledText
from tkMessageBox import showinfo
import tkFileDialog
#jim tracker
from customer import CustomerFrame, Customers
from payment import PaymentFrame, Payments
from log_data import CheckInFrame
from reports import ReportsFrame,generate_info_file
from admin import AdminFrame
#python libraries
from os import listdir, getcwd, chdir
from datetime import datetime
#openpyxl
from openpyxl import load_workbook

class JimNotebook(Frame):
    def __init__(self, name='notebookdemo'):
        Frame.__init__(self, name=name)
        self.pack(expand=True, fill='both')
        self.master.title('Jim Tracker')

        init_text = self.initalize()

        #variables
        self.customers = Customers()
        self.payments = Payments()

        #notebook
        self.nb = Notebook(self, name='notebook')

        #frames
        self.ci_frame = CheckInFrame(self.nb, self.customers, self.payments,self.refresh)
        self.pt_frame = PaymentFrame(self.nb, self.customers, self.payments, self.output_text, self.refresh)
        self.cu_frame = CustomerFrame(self.nb, self.customers, self.output_text, self.refresh)
        self.re_frame = ReportsFrame(self.nb, self.customers, self.payments, self.output_text)
        self.ad_frame = AdminFrame(self.nb,self.customers,self.payments,self.output_text,self.refresh,self.root_dir)

        #add to notebook
        self.nb.add(self.ci_frame, text="Check In")
        self.nb.add(self.cu_frame, text="Customers")
        self.nb.add(self.pt_frame, text="Payments")
        self.nb.add(self.re_frame, text="Reports")
        self.nb.add(self.ad_frame, text="Admin")

        #pack notebook
        self.nb.pack(expand=True,fill='both',side='top')

        #output log
        stf = Frame(self)
        stf.pack(fill='x',side='top')
        self.scrolled_text = ScrolledText(stf,height=10,width=50,wrap='word',state='disabled')
        self.scrolled_text.pack(expand=True,fill='both')

        self.output_text(init_text)

    def output_text(self,outstr):
        self.scrolled_text['state'] = 'normal'
        self.scrolled_text.insert('end',outstr)
        self.scrolled_text.see('end')
        self.scrolled_text['state'] = 'disabled'

    def refresh(self):
        self.ci_frame.update_values() # updates the values in the check in window if open
        self.pt_frame.update_names() # update names in payments drop down boxes
        self.cu_frame.reset_values() # clear out the name value and reset date in new customer
        self.cu_frame.update_names() # update those names in the edit dialog
        self.re_frame.update() # update years and months in report frame
        self.ad_frame.update() # refresh years, months and cwd

    def initalize(self):
        self.root_dir = getcwd()

        files = listdir(getcwd())
        output = "Initializing - " + str(datetime.today()) + '\n'
        if "config.ini" in files:
            try:
                fh = open("config.ini", 'r')
            except IOError:
                showinfo("File Open in another Program", "Please close config.ini and reopen the program.")
                self.destroy()
            directory = fh.readline()
            chdir(directory)
            fh.close()
        else: # first tidme
            try:
                fh = open("config.ini", 'w')
            except IOError:
                showinfo("File Open in another Program", "Please close config.ini and reopen the program.")
                self.destroy()
            directory = ''
            while directory == '':
                showinfo("Initialization", "Please select a directory to store all data files.")
                directory = tkFileDialog.askdirectory(parent=self,initialdir=getcwd())
            chdir(directory)
            fh.write(directory)
            fh.close()

        files = listdir(getcwd()) #get the files list again as cwd probably changed
        if "jim_info.xlsx" in files:
            try:
                wb = load_workbook("jim_info.xlsx")
            except IOError:
                showinfo("File Open in another Program", "Please close jim_info.xlsx and reopen the program.")
                self.destroy()
            sh = wb.get_sheet_by_name("Admin")
            date = sh.cell(row=1,column=1).value
            if date.month < datetime.today().month or date.year < datetime.today().year:
                output += "A - New Month, copy Punch Cards from last month on Admin tab.\n"
            sh.cell(row=1,column=1).value = datetime.today()
            try:
                wb.save("jim_info.xlsx")
            except IOError:
                showinfo("File Open in another Program", "Please close jim_info.xlsx and reopen the program.")
                self.destroy()
        else:
            output += "! - jim_info.xlsx not found in working directory.\n"
            generate_info_file()
            output += "A - Generated empty jim_info.xlsx file into working directory.\n"

        return output

if __name__ == '__main__':
    JimNotebook().mainloop()